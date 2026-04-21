from django.contrib import admin
from django.shortcuts import render, redirect
from django.urls import path
from django.contrib import messages
from django.db import transaction
import csv
import openpyxl
from io import TextIOWrapper
from .models import Faculty, Department, Session, Course, Student, Result

class FacultyAdmin(admin.ModelAdmin):
    list_display = ['name', 'code']
    search_fields = ['name', 'code']

class DepartmentAdmin(admin.ModelAdmin):
    list_display = ['name', 'faculty', 'code']
    list_filter = ['faculty']
    search_fields = ['name', 'code']

class SessionAdmin(admin.ModelAdmin):
    list_display = ['name', 'start_year', 'end_year', 'is_current']
    list_editable = ['is_current']
    
    def save_model(self, request, obj, form, change):
        # If this session is set as current, unset all others
        if obj.is_current:
            Session.objects.exclude(pk=obj.pk).update(is_current=False)
        super().save_model(request, obj, form, change)

class CourseAdmin(admin.ModelAdmin):
    list_display = ['code', 'name', 'department', 'level']
    list_filter = ['department', 'level']
    search_fields = ['code', 'name']
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('bulk-upload/', self.bulk_upload_courses, name='bulk_upload_courses'),
        ]
        return custom_urls + urls
    
    def bulk_upload_courses(self, request):
        if request.method == 'POST':
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please select a file')
                return redirect('..')
            
            try:
                success_count, error_count, errors = self.process_course_file(file)
                
                if success_count > 0:
                    messages.success(request, f'Successfully uploaded {success_count} courses')
                if error_count > 0:
                    for error in errors[:10]:  # Show first 10 errors
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.error(request, f'...and {len(errors) - 10} more errors')
                
                return redirect('..')
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                return redirect('..')
        
        return render(request, 'admin/bulk_upload.html', {
            'title': 'Bulk Upload Courses',
            'sample_format': 'department_code,course_code,course_name,level',
            'sample_data': 'CSC,CSC101,Introduction to Computing,100',
        })
    
    def process_course_file(self, file):
        success_count = 0
        error_count = 0
        errors = []
        
        if file.name.endswith('.csv'):
            decoded_file = TextIOWrapper(file.file, encoding='utf-8')
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif file.name.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            raise ValueError('File must be CSV or Excel format')
        
        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    dept_code = str(row.get('department_code', '')).strip()
                    course_code = str(row.get('course_code', '')).strip()
                    course_name = str(row.get('course_name', '')).strip()
                    level = int(row.get('level', 0))
                    
                    if not all([dept_code, course_code, course_name, level]):
                        errors.append(f'Row {idx}: Missing required fields')
                        error_count += 1
                        continue
                    
                    department = Department.objects.get(code=dept_code)
                    
                    Course.objects.update_or_create(
                        department=department,
                        code=course_code,
                        defaults={
                            'name': course_name,
                            'level': level
                        }
                    )
                    success_count += 1
                except Department.DoesNotExist:
                    errors.append(f'Row {idx}: Department {dept_code} not found')
                    error_count += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
                    error_count += 1
        
        return success_count, error_count, errors
    
    change_list_template = 'admin/course_changelist.html'

class StudentAdmin(admin.ModelAdmin):
    list_display = ['matric_number', 'name', 'department', 'level', 'session_admitted']
    list_filter = ['department', 'level', 'session_admitted']
    search_fields =  ['matric_number', 'name']

    
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('bulk-upload/', self.bulk_upload_students, name='bulk_upload_students'),
            path('promote/', self.promote_students, name='promote_students'),  # ADD THIS
        ]
        return custom_urls + urls
    
    def promote_students(self, request):
        if request.method == 'POST':
            department_id = request.POST.get('department')
            from_level = int(request.POST.get('from_level', 0))
            
            qs = Student.objects.filter(level=from_level)
            if department_id:
                qs = qs.filter(department_id=department_id)
            
            next_level = from_level + 100
            count = qs.update(level=next_level)
            messages.success(request, f'Promoted {count} students from {from_level}L to {next_level}L')
            return redirect('../')
    
        from .models import Department
        return render(request, 'admin/promote_students.html', {
            'departments': Department.objects.all(),
            'level_choices': Student._meta.get_field('level').choices,
        })
    
    def bulk_upload_students(self, request):
        if request.method == 'POST':
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please select a file')
                return redirect('..')
            
            try:
                success_count, error_count, errors = self.process_student_file(file)
                
                if success_count > 0:
                    messages.success(request, f'Successfully uploaded {success_count} students')
                if error_count > 0:
                    for error in errors[:10]:
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.error(request, f'...and {len(errors) - 10} more errors')
                
                return redirect('..')
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                return redirect('..')
        
        return render(request, 'admin/bulk_upload.html', {
            'title': 'Bulk Upload Students',
            'sample_format': 'matric_number,name,department_code,level,session_name',
            'sample_data': 'CSC/2023/001,John Doe,CSC,100,2023/24',
        })
    
    def process_student_file(self, file):
        success_count = 0
        error_count = 0
        errors = []
        
        if file.name.endswith('.csv'):
            decoded_file = TextIOWrapper(file.file, encoding='utf-8')
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif file.name.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            raise ValueError('File must be CSV or Excel format')
        
        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    matric = str(row.get('matric_number', '')).strip()
                    name = str(row.get('name', '')).strip()
                    
                    dept_code = str(row.get('department_code', '')).strip()
                    level = int(row.get('level', 0))
                    session_name = str(row.get('session_name', '')).strip()
                    
                    if not all([matric, name, dept_code, level, session_name]):
                        errors.append(f'Row {idx}: Missing required fields')
                        error_count += 1
                        continue
                    
                    department = Department.objects.get(code=dept_code)
                    session = Session.objects.get(name=session_name)
                    
                    Student.objects.update_or_create(
                        matric_number=matric,
                        defaults={
                            'name': name,
                            
                            'department': department,
                            'level': level,
                            'session_admitted': session
                        }
                    )
                    success_count += 1
                except Department.DoesNotExist:
                    errors.append(f'Row {idx}: Department {dept_code} not found')
                    error_count += 1
                except Session.DoesNotExist:
                    errors.append(f'Row {idx}: Session {session_name} not found')
                    error_count += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
                    error_count += 1
        
        return success_count, error_count, errors
    
    change_list_template = 'admin/student_changelist.html'

class ResultAdmin(admin.ModelAdmin):
    list_display = ['student', 'course', 'session', 'semester', 'test_score', 'exam_score', 'total_score']
    list_filter = ['session', 'semester', 'course__department', 'course__level']
    search_fields = ['student__matric_number', 'student__name', 'course__code']
    readonly_fields = ['total_score', 'created_at', 'updated_at']
    
    def get_queryset(self, request):
        return super().get_queryset(request).select_related('student', 'course', 'session')
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('bulk-upload/', self.bulk_upload_results, name='bulk_upload_results'),
        ]
        return custom_urls + urls
    
    def bulk_upload_results(self, request):
        if request.method == 'POST':
            file = request.FILES.get('file')
            if not file:
                messages.error(request, 'Please select a file')
                return redirect('..')
            
            try:
                success_count, error_count, errors = self.process_result_file(file)
                
                if success_count > 0:
                    messages.success(request, f'Successfully uploaded {success_count} results')
                if error_count > 0:
                    for error in errors[:10]:
                        messages.error(request, error)
                    if len(errors) > 10:
                        messages.error(request, f'...and {len(errors) - 10} more errors')
                
                return redirect('..')
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
                return redirect('..')
        
        return render(request, 'admin/bulk_upload.html', {
            'title': 'Bulk Upload Results',
            'sample_format': 'matric_number,course_code,session_name,semester,test_score,exam_score',
            'sample_data': 'CSC/2023/001,CSC101,2023/24,First,25,60',
        })
    
    def process_result_file(self, file):
        success_count = 0
        error_count = 0
        errors = []
        
        if file.name.endswith('.csv'):
            decoded_file = TextIOWrapper(file.file, encoding='utf-8')
            reader = csv.DictReader(decoded_file)
            rows = list(reader)
        elif file.name.endswith(('.xlsx', '.xls')):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        else:
            raise ValueError('File must be CSV or Excel format')
        
        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    matric = str(row.get('matric_number', '')).strip()
                    course_code = str(row.get('course_code', '')).strip()
                    session_name = str(row.get('session_name', '')).strip()
                    semester = str(row.get('semester', '')).strip()
                    test_score = row.get('test_score')
                    exam_score = row.get('exam_score')
                    
                    if not all([matric, course_code, session_name, semester]):
                        errors.append(f'Row {idx}: Missing required fields')
                        error_count += 1
                        continue
                    
                    # Convert scores to float or None
                    test_score = float(test_score) if test_score not in [None, ''] else None
                    exam_score = float(exam_score) if exam_score not in [None, ''] else None
                    
                    student = Student.objects.get(matric_number=matric)
                    course = Course.objects.get(code=course_code)
                    session = Session.objects.get(name=session_name)
                    
                    Result.objects.update_or_create(
                        student=student,
                        course=course,
                        session=session,
                        semester=semester,
                        defaults={
                            'test_score': test_score,
                            'exam_score': exam_score
                        }
                    )
                    success_count += 1
                except Student.DoesNotExist:
                    errors.append(f'Row {idx}: Student {matric} not found')
                    error_count += 1
                except Course.DoesNotExist:
                    errors.append(f'Row {idx}: Course {course_code} not found')
                    error_count += 1
                except Session.DoesNotExist:
                    errors.append(f'Row {idx}: Session {session_name} not found')
                    error_count += 1
                except Exception as e:
                    errors.append(f'Row {idx}: {str(e)}')
                    error_count += 1
        
        return success_count, error_count, errors
    
    change_list_template = 'admin/result_changelist.html'

admin.site.register(Faculty, FacultyAdmin)
admin.site.register(Department, DepartmentAdmin)
admin.site.register(Session, SessionAdmin)
admin.site.register(Course, CourseAdmin)
admin.site.register(Student, StudentAdmin)
admin.site.register(Result, ResultAdmin)